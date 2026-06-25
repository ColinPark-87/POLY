# -*- coding: utf-8 -*-
# 위례·유성 셔틀 엑셀 → 정규화 JSON ({classtime,hocha,name,arr,dep,loc_arr,loc_dep,note,address})
import openpyxl, re, json
SRC='C:/Users/user/Desktop/Colin 작업폴더/Raw/0616/버스 적용/'
OUT='C:/Users/user/Desktop/Colin 작업폴더/산출물/버스적용_2026-06-16/_분석/'

def cs(c):
    if c is None: return ''
    if hasattr(c,'strftime'):
        try: return c.strftime('%H:%M')
        except: return str(c)
    return str(c).replace('\n',' ').strip()

# "(동, 아파트) 동호수" → (정류장명=아파트, 행정동)
def parse_stop_paren(s):
    s=str(s or '').strip()
    m=re.match(r'\(([^)]*)\)\s*(.*)', s)
    if m:
        inner=m.group(1)
        parts=[p.strip() for p in inner.split(',')]
        dong = parts[0] if parts and re.search(r'동$', parts[0]) else ''
        apt = parts[1] if len(parts)>1 else (parts[0] if not dong else '')
        return (apt or inner).strip(), dong
    return s, ''

def classify(text):
    t=text
    if re.search(r'유치',t): return '유치부'
    if re.search(r'매일|5일',t): return '매일반'
    if re.search(r'월[·.]?수[·.]?금|월수금|3일',t): return '월수금'
    if re.search(r'화[·.]?목|화목|2일',t): return '화목'
    return None

# ── 위례: 호차별 시트, ◼ 섹션 ──────────────────────────────
def parse_wirye():
    wb=openpyxl.load_workbook(SRC+'위례_ 2026년 셔틀 전 호차.xlsx',data_only=True)
    recs=[]
    for sh in wb.sheetnames:
        m=re.match(r'(\d+)호차',sh)
        if not m: continue
        hocha=m.group(1)+'호차'
        ws=wb[sh]; rows=[[cs(c) for c in r] for r in ws.iter_rows(values_only=True)]
        cur=None; incols=None
        for i,r in enumerate(rows):
            line=' '.join(r)
            if '◼' in line or re.match(r'^\s*[◼■◾▪]',line):
                cl=classify(line);
                if cl: cur=cl; continue
            if 'No.' in r or ('탑승' in line and '이름' in line):
                # 컬럼 인덱스
                incols={'no':None,'loc':None,'arr':None,'dep':None,'name':None,'lv':None,'note':None}
                for j,h in enumerate(r):
                    if h.startswith('No'): incols['no']=j
                    elif '탑승' in h: incols['loc']=j
                    elif '등원' in h: incols['arr']=j
                    elif '하원' in h: incols['dep']=j
                    elif h=='이름': incols['name']=j
                    elif '학급' in h: incols['lv']=j
                    elif '비고' in h: incols['note']=j
                continue
            if not cur or not incols or incols['name'] is None: continue
            nm=r[incols['name']] if incols['name']<len(r) else ''
            if not nm or '합계' in line or nm=='이름': continue
            rawloc=r[incols['loc']] if incols['loc'] is not None and incols['loc']<len(r) else ''
            stop,dong=parse_stop_paren(rawloc)
            recs.append({'classtime':cur,'hocha':hocha,'name':nm,
                'arr':r[incols['arr']] if incols['arr']<len(r) else '',
                'dep':r[incols['dep']] if incols['dep']<len(r) else '',
                'loc_arr':stop,'loc_dep':stop,'address':rawloc,'dong':dong,
                'note':r[incols['note']] if incols['note'] is not None and incols['note']<len(r) else ''})
    wb.close()
    return recs

# ── 유성: 호차별 시트, 세션라인 ───────────────────────────
def parse_yuseong():
    wb=openpyxl.load_workbook(SRC+'유성_셔틀명단 260615.xlsx',data_only=True)
    recs=[]
    for sh in wb.sheetnames:
        m=re.match(r'(\d+)호차',sh)
        if not m: continue
        hocha=m.group(1)+'호차'
        ws=wb[sh]; rows=[[cs(c) for c in r] for r in ws.iter_rows(values_only=True)]
        cur=None; prev_loc=''; prev_arr=''; prev_dep=''
        for r in rows:
            line=' '.join(r); nonblank=[x for x in r if x]
            # 세션 마커
            if len(nonblank)<=3 and classify(line) and ':' in line or (len(nonblank)<=2 and classify(line)):
                cl=classify(line)
                if cl: cur=cl; prev_loc=prev_arr=prev_dep=''; continue
            if 'NO' in r or '승차장소' in line: prev_loc=prev_arr=prev_dep=''; continue
            if not cur: continue
            # 컬럼: NO|승차장소|이름|전화|등원|하원|비고
            if len(r)<6: continue
            nm=r[2]
            if not nm or nm=='이름': continue
            # 정류장 빈칸 = 윗행과 같은 정류장 → 시간도 상속(같은 정류장 같은 픽업)
            same_stop = not r[1]
            loc = r[1] or prev_loc
            arr = r[4] or (prev_arr if same_stop else '')
            dep = r[5] or (prev_dep if same_stop else '')
            if r[1]: prev_loc=r[1]; prev_arr=r[4]; prev_dep=r[5]
            recs.append({'classtime':cur,'hocha':hocha,'name':nm,
                'arr':arr,'dep':dep,'loc_arr':loc,'loc_dep':loc,'address':loc,
                'note':r[6] if len(r)>6 else ''})
    wb.close()
    return recs

for name,fn in [('위례',parse_wirye),('유성',parse_yuseong)]:
    recs=fn()
    json.dump(recs,open(OUT+name+'_parsed.json','w',encoding='utf-8'),ensure_ascii=False,indent=1)
    from collections import Counter
    print(f"{name}: {len(recs)}행 | classtime={dict(Counter(r['classtime'] for r in recs))} | hocha={dict(Counter(r['hocha'] for r in recs))}")
    print("  샘플:",json.dumps(recs[0],ensure_ascii=False) if recs else 'none')
