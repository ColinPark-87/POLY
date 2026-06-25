# -*- coding: utf-8 -*-
# 운정 셔틀 → 정규화 JSON. 세션×호차 시트(매일반은 등원/하원 분리). 라벨로 컬럼 탐지.
import openpyxl, re, json
SRC='C:/Users/user/Desktop/Colin 작업폴더/Raw/0616/버스 적용/'
OUT='C:/Users/user/Desktop/Colin 작업폴더/산출물/버스적용_2026-06-16/_분석/'
def cs(c):
    if c is None: return ''
    if hasattr(c,'strftime'):
        try: return c.strftime('%H:%M')
        except: return str(c)
    return str(c).replace('\n',' ').strip()
def classify(t):
    if re.search(r'유치',t): return '유치부'
    if re.search(r'매일|5일',t): return '매일반'
    if re.search(r'월수금|3일',t): return '월수금'
    if re.search(r'화목|2일',t): return '화목'
    return None
def parse_stop_paren(s):
    s=str(s or '').strip()
    m=re.match(r'.*?\(([^)]*)\)\s*(.*)', s)
    if m:
        parts=[p.strip() for p in m.group(1).split(',')]
        dong=parts[0] if parts and re.search(r'동$',parts[0]) else ''
        apt=parts[1] if len(parts)>1 else (parts[0] if not dong else '')
        return (apt or m.group(1)).strip(), dong
    return s, ''

wb=openpyxl.load_workbook(SRC+'운정_2026 전체 셔틀명단_(6월).xlsx',data_only=True)
recs=[]
for sh in wb.sheetnames:
    cl=classify(sh); hm=re.search(r'(\d+)호[차처]',sh)
    if not cl or not hm: continue
    hocha=hm.group(1)+'호차'
    ws=wb[sh]; rows=[[cs(c) for c in r] for r in ws.iter_rows(values_only=True)]
    # 헤더존(상위 6행)에서 라벨→컬럼
    col={'arr':None,'dep':None,'loc':None,'name':None}; hend=0
    for i,r in enumerate(rows[:7]):
        for j,h in enumerate(r):
            if h=='등원' and col['arr'] is None: col['arr']=j
            elif h=='하원' and col['dep'] is None: col['dep']=j
            elif '출발장소' in h: col['loc']=j
            elif h=='이름': col['name']=j
        if 'No.' in r or 'No' in r or '이름' in r: hend=i
    if col['name'] is None: continue
    for r in rows[hend+1:]:
        nm=(r[col['name']] if col['name']<len(r) else '').lstrip('/ ').strip()
        if not nm or '이름' in nm: continue
        loc=r[col['loc']] if col['loc'] is not None and col['loc']<len(r) else ''
        stop,dong=parse_stop_paren(loc)
        arr=r[col['arr']] if col['arr'] is not None and col['arr']<len(r) else ''
        dep=r[col['dep']] if col['dep'] is not None and col['dep']<len(r) else ''
        recs.append({'classtime':cl,'hocha':hocha,'name':nm,'arr':arr,'dep':dep,
            'loc_arr':stop,'loc_dep':stop,'address':loc,'dong':dong,
            'note':''})
wb.close()
json.dump(recs,open(OUT+'운정_parsed.json','w',encoding='utf-8'),ensure_ascii=False,indent=1)
from collections import Counter
print(f"운정: {len(recs)}행 | classtime={dict(Counter(r['classtime'] for r in recs))} | hocha={dict(Counter(r['hocha'] for r in recs))}")
print("샘플(유치):",json.dumps([r for r in recs if r['classtime']=='유치부'][0],ensure_ascii=False))
print("샘플(매일):",json.dumps([r for r in recs if r['classtime']=='매일반'][0],ensure_ascii=False))
# 호차 목록
buses=sorted({r['hocha'] for r in recs},key=lambda x:int(x.replace('호차','')))
json.dump([{'name':b} for b in buses],open(OUT+'운정_buses.json','w',encoding='utf-8'),ensure_ascii=False,indent=1)
print('호차:',buses)
