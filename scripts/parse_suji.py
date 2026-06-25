# -*- coding: utf-8 -*-
# 수지 셔틀 → 정규화 JSON. 각 호차 시트, col4="{호차} {세션} {등원}/{하원} {등원정류장}/{하원정류장}"
import openpyxl, re, json
SRC='C:/Users/user/Desktop/Colin 작업폴더/Raw/0616/버스 적용/'
OUT='C:/Users/user/Desktop/Colin 작업폴더/산출물/버스적용_2026-06-16/_분석/'
def cs(c):
    if c is None: return ''
    if hasattr(c,'strftime'):
        try: return c.strftime('%H:%M')
        except: return str(c)
    return str(c).replace('\n',' ').strip()
def classify(t):
    if re.search(r'유[치지]부',t): return '유치부'
    if re.search(r'매일|5일',t): return '매일반'
    if re.search(r'월수금|3일',t): return '월수금'
    if re.search(r'화목|2일',t): return '화목'
    return None
def dong_of(addr):
    m=re.search(r'([가-힣]+동)\b',str(addr or ''))
    return m.group(1) if m else ''

wb=openpyxl.load_workbook(SRC+'수지캠퍼스_26년 6월 셔틀운행정보_20260615.xlsx',data_only=True)
recs=[]
for sh in wb.sheetnames:
    m=re.match(r'(\d+)호',sh)
    if not m: continue  # 셔틀안탐/유치부전체 제외
    hocha=m.group(1)+'호차'
    ws=wb[sh]
    for r in ws.iter_rows(values_only=True):
        v=[cs(c) for c in r]
        if len(v)<5: continue
        nm=v[2] if len(v)>2 else ''
        info=v[4] if len(v)>4 else ''
        if not nm or nm=='이름' or '셔틀표' in ' '.join(v): continue
        cl=classify(info)
        if not cl: continue
        # col4 파싱: 호차 세션 등원/하원 [정류장/정류장]
        mm=re.search(r'(\d{1,2}:\d{2})\s*/\s*(\d{1,2}:\d{2})\s*(.*)', info)
        arr=dep=''; loc_arr=loc_dep=''
        if mm:
            arr=mm.group(1); dep=mm.group(2); stops=mm.group(3).strip()
            parts=[p.strip() for p in stops.split('/') if p.strip()]
            if len(parts)>=2: loc_arr,loc_dep=parts[0],parts[1]
            elif len(parts)==1: loc_arr=loc_dep=parts[0]
        addr=v[5] if len(v)>5 else ''
        recs.append({'classtime':cl,'hocha':hocha,'name':nm,'arr':arr,'dep':dep,
            'loc_arr':loc_arr,'loc_dep':loc_dep,'address':addr,'dong':dong_of(addr),
            'note':v[6] if len(v)>6 else ''})
wb.close()
json.dump(recs,open(OUT+'수지_parsed.json','w',encoding='utf-8'),ensure_ascii=False,indent=1)
from collections import Counter
print(f"수지: {len(recs)}행 | classtime={dict(Counter(r['classtime'] for r in recs))} | hocha={dict(Counter(r['hocha'] for r in recs))}")
print("샘플:",json.dumps(recs[0],ensure_ascii=False))
print("샘플2:",json.dumps([r for r in recs if r['classtime']!='유치부'][0],ensure_ascii=False))
